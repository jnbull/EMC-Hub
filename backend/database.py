import openpyxl
from models import addEquipment

wb = openpyxl.load_workbook(filename='./assets/files/Equipment.xlsx')
ws = wb._sheets[0]

firstCellValue = 0
row = 9

while firstCellValue is not None:
    id = ws.cell(column=6, row=row).value
    name = ws.cell(column=2, row=row).value
    type_ = ws.cell(column=1, row=row).value
    manufacturer = ws.cell(column=3, row=row).value
    calDate = ws.cell(column=4, row=row).value
    calDue = ws.cell(column=5, row=row).value
    firstCellValue = ws.cell(column=1, row=row).value
    row += 1

    if id is not None:
        addEquipment(id, name, type_, manufacturer, calDate, calDue)
